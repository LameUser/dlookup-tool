import os
import shutil
import time
import threading
import concurrent.futures
import openpyxl
from subprocess import run, PIPE, CalledProcessError
from tldextract import extract
import requests
import pyfiglet
from termcolor import colored


completed_tasks = 0  # Counter for processed domains
total_tasks = 0  # Total number of domains
screenshot_tasks = 0  # Counter for completed screenshots
total_screenshots = 0  # Total number of screenshots


def display_banner():
    """Display the tool banner with the name and GitHub link."""
    banner = pyfiglet.figlet_format("dlookup", font="slant")  # Create banner text
    print(colored(banner, "cyan"))  # Print banner in cyan
    print(colored("https://github.com/LameUser", "yellow"))  # Print GitHub link in yellow


def clean_domain(domain):
    """Extract main domain using tldextract."""
    extracted = extract(domain)
    return f"{extracted.domain}.{extracted.suffix}" if extracted.domain and extracted.suffix else None


def run_command(command):
    """Run a shell command and handle errors."""
    try:
        result = run(command, shell=True, text=True, stdout=PIPE, stderr=PIPE, timeout=10)
        if result.returncode == 0:
            return result.stdout.strip()
        else:
            return f"Command failed: {result.stderr.strip()}"
    except Exception as e:
        return f"Error running command '{command}': {e}"


def extract_server_ip(nslookup_result):
    """Extract the server IP from NSLOOKUP results."""
    lines = nslookup_result.splitlines()
    ip_addresses = [line.split(":")[-1].strip() for line in lines if "Address:" in line and not line.startswith("Server:")]
    return ip_addresses[-1] if ip_addresses else "N/A"


def get_geolocation(ip):
    """Fetch geolocation data for a given IP address using ip-api.com."""
    if ip == "N/A" or not ip:  # Skip if IP is invalid
        return {'Country': 'N/A', 'City': 'N/A', 'ISP': 'N/A', 'ASN': 'N/A'}

    url = f"http://ip-api.com/json/{ip}"
    retries = 3
    for attempt in range(retries):
        try:
            response = requests.get(url, timeout=5)
            if response.status_code == 200:
                data = response.json()
                if data.get('status') == 'success':
                    return {
                        'Country': data.get('country', 'N/A'),
                        'City': data.get('city', 'N/A'),
                        'ISP': data.get('isp', 'N/A'),
                        'ASN': data.get('as', 'N/A').split()[0]  # Extract ASN
                    }
            time.sleep(2)  # Add a delay between retries
        except:
            time.sleep(2)  # Add a delay for failed attempts
    return {'Country': 'N/A', 'City': 'N/A', 'ISP': 'N/A', 'ASN': 'N/A'}


def process_domain(domain):
    """Process a single domain."""
    global completed_tasks
    main_domain = clean_domain(domain)
    if not main_domain:
        completed_tasks += 1
        return domain, None, "Invalid Domain", "N/A", "N/A", "No", None, "N/A", "N/A", "N/A", "N/A"

    # WHOIS command
    whois_result = run_command(f"whois {main_domain}")

    # NSLOOKUP command
    nslookup_result = run_command(f"nslookup {main_domain}")

    # Extract Server IP from NSLOOKUP result
    server_ip = extract_server_ip(nslookup_result)

    # Check domain activity (using IP as an indicator)
    domain_active = "Yes" if server_ip != "N/A" else "No"

    # Geolocation lookup
    geo_data = get_geolocation(server_ip)

    completed_tasks += 1
    return (
        domain,
        main_domain,
        whois_result,
        nslookup_result,
        server_ip,
        domain_active,
        None,  # Successful Scheme (HTTP/HTTPS, if applicable)
        geo_data['Country'],
        geo_data['City'],
        geo_data['ISP'],
        geo_data['ASN']
    )


def display_progress():
    """Manually display progress when user hits Enter."""
    global completed_tasks, total_tasks
    while completed_tasks < total_tasks:
        input("\nPress Enter to see domain processing progress... ")
        print(f"Progress: {completed_tasks}/{total_tasks} tasks completed "
              f"({(completed_tasks / total_tasks) * 100:.2f}%).")


def capture_screenshots(txt_file, screenshot_dir):
    """Run EyeWitness to capture screenshots with retries and manual progress."""
    global screenshot_tasks, total_screenshots

    if os.path.exists(screenshot_dir):
        user_choice = input(f"\nOutput directory '{screenshot_dir}' already exists. Do you want to delete it and continue? [y/n]: ").strip().lower()
        if user_choice == "y":
            shutil.rmtree(screenshot_dir)
        else:
            print("Aborting screenshot capture.")
            return

    os.makedirs(screenshot_dir)

    with open(txt_file, 'r') as file:
        urls = file.readlines()

    total_screenshots = len(urls)

    while screenshot_tasks < total_screenshots:
        input("\nPress Enter to see screenshot progress... ")
        print(f"Screenshot Progress: {screenshot_tasks}/{total_screenshots} screenshots completed "
              f"({(screenshot_tasks / total_screenshots) * 100:.2f}%).")

    for url in urls:
        retries = 3
        for attempt in range(retries):
            command = f"eyewitness -f {txt_file} -d {screenshot_dir}"
            try:
                run(command, shell=True, check=True)
                screenshot_tasks += 1
                break
            except CalledProcessError:
                time.sleep(2)  # Wait before retrying
        else:
            screenshot_tasks += 1  # Count it as failed but progress continues


def process_domains(input_file, txt_output, result_output, screenshot_dir):
    """Process all domains and save results to files."""
    global total_tasks
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active
    result_wb = openpyxl.Workbook()
    result_sheet = result_wb.active
    result_sheet.append([
        "URLS", "Domain", "WHOIS Result", "NSLOOKUP Result", "Server IP",
        "Domain Active", "Successful Scheme", "Country", "City", "ISP", "ASN"
    ])

    domains = [row[0].value for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1) if row[0].value]
    total_tasks = len(domains)

    progress_thread = threading.Thread(target=display_progress, daemon=True)
    progress_thread.start()

    active_domains = []

    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(process_domain, domain): domain for domain in domains}
        for future in concurrent.futures.as_completed(futures):
            result = future.result()
            result_sheet.append(result[:11])
            if result[5] == "Yes":
                active_domains.append(f"http://{result[1]}")

    result_wb.save(result_output)
    print(f"Results saved to Excel file: {result_output}")

    with open(txt_output, 'w') as txt_file:
        txt_file.writelines([f"{domain}\n" for domain in active_domains])
    print(f"Active domains saved to text file: {txt_output}")

    if active_domains:
        capture_screenshots(txt_output, screenshot_dir)


if __name__ == "__main__":
    display_banner()

    base_path = input("\nEnter the full path to the folder containing the Excel file: ").strip()
    input_file = os.path.join(base_path, "cryptodomain.xlsx")
    txt_output = os.path.join(base_path, "urls.txt")
    result_output = os.path.join(base_path, "domain_results.xlsx")
    screenshot_dir = os.path.join(base_path, "screens/")

    if not os.path.exists(input_file):
        print(f"Input file not found: {input_file}")
    else:
        process_domains(input_file, txt_output, result_output, screenshot_dir)
